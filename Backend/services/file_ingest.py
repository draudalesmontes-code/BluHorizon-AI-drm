from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi import HTTPException, UploadFile
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


TEXT_EXTENSIONS = {
    ".txt", ".md", ".json", ".py", ".java", ".csv", ".yaml", ".yml", ".xml"
}


async def extract_text_from_upload(file: UploadFile) -> tuple[str, str]:
    filename = file.filename or "uploaded_file"
    ext = Path(filename).suffix.lower()
    data = await file.read()

    if ext in TEXT_EXTENSIONS:
        return data.decode("utf-8", errors="ignore"), ext

    if ext == ".pdf":
        reader = PdfReader(BytesIO(data))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        return text, ext

    if ext == ".docx":
        document = Document(BytesIO(data))
        text = "\n".join(p.text for p in document.paragraphs)
        return text, ext

    if ext == ".xlsx":
        workbook = load_workbook(BytesIO(data), data_only=True)
        lines: list[str] = []

        for sheet in workbook.worksheets:
            lines.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join("" if cell is None else str(cell) for cell in row)
                if row_text.strip():
                    lines.append(row_text)

        return "\n".join(lines), ext

    raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")
